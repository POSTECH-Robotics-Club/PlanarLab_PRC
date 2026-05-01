"""
Extracts a track path from colored Excel cells, orders it, scales it to real-world dimensions,
interpolates points at a fixed resolution, and exports it as a CSV file.
"""

import sys
import csv
from openpyxl import load_workbook
import math

# Neighboring directions (4-connected + diagonals)
DIRS = [
    (0,1),(0,-1),(1,0),(-1,0),  # up, down, left, right
    (1,1),(1,-1),(-1,1),(-1,-1)  # diagonals
]

def read_track_from_excel(xlsx_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet.")

    track_cells = set()
    for row in ws.iter_rows():
        for cell in row:
            # Check if cell has a non-default fill color
            if cell.fill and cell.fill.start_color.type != 'auto' and cell.fill.start_color.rgb != "00000000":  # type: ignore
                track_cells.add((cell.column, cell.row))  # (x, y)
    return track_cells

def find_neighbors(cell, track_cells):
    x, y = cell
    neighbors = []
    for dx, dy in DIRS:
        nxt = (x+dx, y+dy)
        if nxt in track_cells:
            neighbors.append(nxt)
    return neighbors

def order_track(track_cells, closed):
    if closed:
        # For closed tracks, start from the top-left-most cell
        start = min(track_cells, key=lambda c: (c[1], c[0]))
    else:
        # For open tracks, start from an endpoint (degree == 1), preferring top-left
        ends = [c for c in track_cells if len(find_neighbors(c, track_cells)) == 1]
        if not ends:
            raise ValueError("Open track must have endpoints.")
        start = min(ends, key=lambda c: (c[1], c[0]))

    ordered = [start]
    visited = {start}
    current = start

    while True:
        neighbors = [n for n in find_neighbors(current, track_cells) if n not in visited]
        if not neighbors:
            break
        nxt = neighbors[0]
        ordered.append(nxt)
        visited.add(nxt)
        current = nxt

    return ordered

def interpolate_points(ordered_track, scale_x, scale_y, min_x, min_y, resolution):
    result = []
    for i in range(len(ordered_track)-1):
        x1, y1 = ordered_track[i]
        x2, y2 = ordered_track[i+1]
        
        # Convert grid coordinates to real-world coordinates
        x1_m = (x1 - min_x) * scale_x
        y1_m = (y1 - min_y) * scale_y
        x2_m = (x2 - min_x) * scale_x
        y2_m = (y2 - min_y) * scale_y

        result.append((x1_m, y1_m))

        dx = x2_m - x1_m
        dy = y2_m - y1_m
        dist = math.hypot(dx, dy)

        # Interpolate intermediate points if segment is longer than resolution
        if dist > resolution:
            steps = int(dist // resolution)
            for s in range(1, steps+1):
                t = (s * resolution) / dist
                if t < 1:
                    xi = x1_m + dx * t
                    yi = y1_m + dy * t
                    result.append((xi, yi))

    # Append final point
    x_end = (ordered_track[-1][0] - min_x) * scale_x
    y_end = (ordered_track[-1][1] - min_y) * scale_y
    result.append((x_end, y_end))

    return result

def main():
    if len(sys.argv) < 7:
        print("Usage: python track_extract.py <input.xlsx> <output.csv> <open|closed> <width> <height> <resolution>")
        sys.exit(1)

    xlsx_file = sys.argv[1]
    csv_file = sys.argv[2]
    mode = sys.argv[3].lower()
    width = float(sys.argv[4])
    height = float(sys.argv[5])
    resolution = float(sys.argv[6])

    closed = (mode == "closed")

    track_cells = read_track_from_excel(xlsx_file)
    ordered_track = order_track(track_cells, closed)

    # Compute bounding box of grid
    max_x = max(x for x, _ in track_cells)
    min_x = min(x for x, _ in track_cells)
    max_y = max(y for _, y in track_cells)
    min_y = min(y for _, y in track_cells)

    grid_w = max_x - min_x + 1
    grid_h = max_y - min_y + 1

    # Scale factors to real-world dimensions
    scale_x = width / grid_w
    scale_y = height / grid_h

    interpolated_track = interpolate_points(
        ordered_track, scale_x, scale_y, min_x, min_y, resolution
    )

    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["x_m", "y_m"])
        for x_m, y_m in interpolated_track:
            writer.writerow([x_m, y_m])

if __name__ == "__main__":
    main()